"""
Smart encoding fix: 
1. Use geo columns (28=Estago_geo, 30=Municipio_geo) which have CORRECT encoding
2. Build a token-level fix map by comparing broken municipio (col 3) with correct Municipio_geo (col 30)
3. Apply those token fixes to localidades and market names
"""
import openpyxl, sys, io, re
from collections import defaultdict
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('mercados_estado_mun.xlsx')
ws = wb['capa_unida']

# Step 1: Build token fix map from municipio vs Municipio_geo comparison
token_fixes = {}  # broken_token → correct_token
full_fixes = {}   # broken_full_string → correct_full_string (for multi-word validation)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    mun = row[3]     # municipio (broken)
    mun_geo = row[30] # Municipio_geo (correct)
    
    if mun and mun_geo and '\ufffd' in str(mun):
        mun_s = str(mun)
        mun_g = str(mun_geo)
        full_fixes[mun_s] = mun_g
        
        # Also extract word-level tokens
        broken_words = mun_s.split()
        correct_words = mun_g.split()
        
        if len(broken_words) == len(correct_words):
            for bw, cw in zip(broken_words, correct_words):
                if '\ufffd' in bw and bw != cw:
                    token_fixes[bw] = cw

# Also add entidad fixes from col 2 vs col 28
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    ent = row[2]
    ent_geo = row[28]
    if ent and ent_geo and '\ufffd' in str(ent):
        ent_s = str(ent)
        ent_g = str(ent_geo)
        full_fixes[ent_s] = ent_g
        broken_words = ent_s.split()
        correct_words = ent_g.split()
        if len(broken_words) == len(correct_words):
            for bw, cw in zip(broken_words, correct_words):
                if '\ufffd' in bw and bw != cw:
                    token_fixes[bw] = cw

print(f"=== TOKEN FIXES FROM GEO COLUMNS ({len(token_fixes)}) ===")
for broken, correct in sorted(token_fixes.items()):
    print(f"  {repr(broken):40s} → {correct}")

print(f"\n=== FULL STRING FIXES ({len(full_fixes)}) ===")
for broken, correct in sorted(full_fixes.items()):
    print(f"  {repr(broken):50s} → {correct}")

# Step 2: Now find which broken localidade/market tokens are NOT covered
unique_markets = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    mid = row[13]
    if mid and mid not in unique_markets:
        unique_markets[mid] = row

uncovered_tokens = set()
uncovered_full = set()

for mid, r in unique_markets.items():
    for idx in [4, 14]:  # localidad, market_name
        val = r[idx]
        if val and '\ufffd' in str(val):
            s = str(val)
            # Try token-level fix
            words = s.split()
            still_broken = False
            for w in words:
                if '\ufffd' in w and w not in token_fixes:
                    # Check if it's a known token with parentheses/brackets
                    clean = w.strip('()[]')
                    if clean in token_fixes:
                        continue
                    uncovered_tokens.add(w)
                    still_broken = True
            if still_broken:
                uncovered_full.add(s)

print(f"\n=== UNCOVERED TOKENS ({len(uncovered_tokens)}) ===")
for t in sorted(uncovered_tokens):
    print(f"  {repr(t)}")

print(f"\n=== STRINGS WITH UNCOVERED TOKENS ({len(uncovered_full)}) ===")
for s in sorted(uncovered_full):
    print(f"  {repr(s)}")
