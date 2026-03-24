import openpyxl

wb = openpyxl.load_workbook('mercados_estado_mun.xlsx')
ws = wb['capa_unida']

# Mapa de correcciones de encoding roto -> texto correcto
FIX = {
    'Ciudad de M\ufffdxico': 'Ciudad de México',
    'M\ufffdxico': 'México',
    'Michoac\ufffdn de Ocampo': 'Michoacán de Ocampo',
    'Nuevo Le\ufffdn': 'Nuevo León',
    'Quer\ufffdtaro': 'Querétaro',
    'San Luis Potos\ufffd': 'San Luis Potosí',
    'Yucat\ufffdn': 'Yucatán',
}

def fix_text(txt):
    if txt is None:
        return None
    s = str(txt)
    for bad, good in FIX.items():
        s = s.replace(bad, good)
    # Fix individual replacement chars remaining
    # Common patterns: \ufffd = replacement char
    # We'll also do a broad scan for known municipality/localidad names
    return s

# Build a broader fix map by scanning what's broken
broken_chars = {}
catalogo_limpio = {}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    estatus = row[16]
    market_id = row[13]
    if market_id and estatus == 'CATALOGO_LIMPIO':
        if market_id not in catalogo_limpio:
            catalogo_limpio[market_id] = {
                'market_id': market_id,
                'market_name': fix_text(row[14]),
                'tipo': row[15],
                'entidad': fix_text(row[2]),
                'municipio': fix_text(row[3]),
                'localidad': fix_text(row[4]),
                'latitud': row[5],
                'longitud': row[6],
                'n_establecimientos': row[18],
                'cve_ent': row[27],
                'cve_mun': row[29],
            }

# Check for remaining replacement chars
remaining = set()
for mk in catalogo_limpio.values():
    for field in ['market_name', 'entidad', 'municipio', 'localidad']:
        val = mk[field]
        if val and '\ufffd' in val:
            remaining.add((field, val))

if remaining:
    print("=== REMAINING BROKEN STRINGS ===")
    for field, val in sorted(remaining):
        print(f"  {field}: {repr(val)}")
    print()

# Print corrected entidades
print("=== CORRECTED ENTIDADES ===")
entidades = set()
for mk in catalogo_limpio.values():
    entidades.add(mk['entidad'])
for e in sorted(entidades):
    print(f"  {e}")

print(f"\nTotal markets: {len(catalogo_limpio)}")
