"""
Check if geo columns have correct encoding and can be used as reference.
Also try to read the raw XML from the xlsx to see if bytes are recoverable.
"""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('mercados_estado_mun.xlsx')
ws = wb['capa_unida']

# Check columns 28 (Estago_geo) and 30 (Municipio_geo) 
print("=== Check geo columns for correct encoding ===")
sample_rows = []
for row in ws.iter_rows(min_row=2, max_row=20, values_only=True):
    sample_rows.append(row)

for r in sample_rows[:10]:
    ent = r[2]  # entidad
    est_geo = r[28]  # Estago_geo
    mun = r[3]  # municipio
    mun_geo = r[30]  # Municipio_geo
    print(f"  entidad={repr(str(ent)[:30])}  Estago_geo={repr(str(est_geo)[:30])}")
    print(f"  municipio={repr(str(mun)[:30])}  Municipio_geo={repr(str(mun_geo)[:30])}")
    print()

# Now check if Estago_geo/Municipio_geo have broken chars
broken_geo = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    for idx in [28, 30]:
        val = row[idx]
        if val and '\ufffd' in str(val):
            broken_geo += 1

print(f"Broken chars in geo columns: {broken_geo}")

# Check: can we read the xlsx as a zip and extract the raw XML?
import zipfile
with zipfile.ZipFile('mercados_estado_mun.xlsx', 'r') as z:
    # List shared strings file
    for name in z.namelist():
        if 'shared' in name.lower():
            print(f"\nFound: {name}")
            data = z.read(name)
            # Check if the raw bytes have the proper encoding
            # Look for a known broken string like "México"
            # In UTF-8, é is 0xC3 0xA9
            # Search for "M" followed by something then "xico"
            idx_start = data.find(b'xico')
            if idx_start > 0:
                snippet = data[idx_start-5:idx_start+10]
                print(f"  Raw bytes near 'xico': {snippet}")
                print(f"  Hex: {snippet.hex()}")
            
            # Search for the replacement char in UTF-8 (EF BF BD)
            replacement_utf8 = b'\xef\xbf\xbd'
            count = data.count(replacement_utf8)
            print(f"  U+FFFD occurrences in raw XML: {count}")
            
            # Check if there are any raw latin-1 bytes instead
            # é in latin-1 is 0xE9
            # Look for patterns
            # Let's look at what's around "M" before "xico"
            pos = 0
            found = 0
            while found < 3:
                pos = data.find(b'xico', pos+1)
                if pos < 0:
                    break
                context = data[max(0,pos-10):pos+10]
                print(f"  Context around 'xico' at {pos}: {context}")
                print(f"  Hex: {context.hex()}")
                found += 1
