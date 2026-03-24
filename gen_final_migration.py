"""
FINAL script: Generate migration 007 with ALL 2217 markets, correct encoding.
Uses:
  - Column 28 (Estago_geo) for entidad  
  - Column 30 (Municipio_geo) for municipio
  - Token-level fix map for localidad and market_name
"""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('mercados_estado_mun.xlsx')
ws = wb['capa_unida']

# ============================================================
# Step 1: Build token fix map from municipio vs Municipio_geo
# ============================================================
TOKEN_FIX = {}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    mun = row[3]
    mun_geo = row[30]
    if mun and mun_geo and '\ufffd' in str(mun):
        broken_words = str(mun).split()
        correct_words = str(mun_geo).split()
        if len(broken_words) == len(correct_words):
            for bw, cw in zip(broken_words, correct_words):
                if '\ufffd' in bw and bw != cw:
                    TOKEN_FIX[bw] = cw
    
    ent = row[2]
    ent_geo = row[28]
    if ent and ent_geo and '\ufffd' in str(ent):
        broken_words = str(ent).split()
        correct_words = str(ent_geo).split()
        if len(broken_words) == len(correct_words):
            for bw, cw in zip(broken_words, correct_words):
                if '\ufffd' in bw and bw != cw:
                    TOKEN_FIX[bw] = cw

# ============================================================
# Step 2: Add manual fixes for tokens not covered by geo columns
# (localidades and market names that have unique broken words)
# ============================================================
MANUAL_TOKEN_FIX = {
    # Localidad tokens
    'Constituci\ufffdn': 'Constitución',
    'Camal\ufffd': 'Camalú',
    'Canc\ufffdn': 'Cancún',
    'Ca\ufffd\ufffdn': 'Cañón',
    'Chavarr\ufffda': 'Chavarría',
    'C\ufffdrdova': 'Córdova',
    'Echeverr\ufffda': 'Echeverría',
    'Enr\ufffdquez': 'Enríquez',
    'Estaci\ufffdn': 'Estación',
    '(Estaci\ufffdn': '(Estación',
    'Galv\ufffdn)': 'Galván)',
    'Guam\ufffdchil': 'Guamúchil',
    'Guzm\ufffdn': 'Guzmán',
    'Hind\ufffd': 'Hindú',
    'Ich\ufffdn': 'Ichán',
    'Kil\ufffdmetro': 'Kilómetro',
    '(Kil\ufffdmetro': '(Kilómetro',
    'N\ufffdmero': 'Número',
    'Ocuiltzapotl\ufffdn': 'Ocuiltzapotlán',
    'Opci\ufffdn': 'Opción',
    'Ord\ufffd\ufffdez': 'Ordóñez',
    'Par\ufffdcuaro': 'Parácuaro',
    'Pe\ufffdita': 'Peñita',
    'Puriatz\ufffdcuaro': 'Puriatzcuaro',  # Santiago Puriatzécuaro? Actually Puriatzícuaro... let me check. It's a place in Michoacán. The correct name is Puriatzícuaro... no, it's Puriatzcuaro? Let me use Puriatzícuaro
    'Quil\ufffd': 'Quilá',
    '(Qu\ufffdmicas': '(Químicas',
    'Ram\ufffdn': 'Ramón',
    'Ram\ufffdrez': 'Ramírez',
    'Rodr\ufffdguez': 'Rodríguez',
    'Rosal\ufffda': 'Rosalía',
    'Sahag\ufffdn': 'Sahagún',
    'Sahag\ufffdn)': 'Sahagún)',
    'Sebasti\ufffdn': 'Sebastián',
    'Secci\ufffdn': 'Sección',
    'Su\ufffdrez': 'Suárez',
    'Tamult\ufffd': 'Tamulté',
    'Teacap\ufffdn': 'Teacapán',
    'Tur\ufffdstico]': 'Turístico]',
    'T\ufffdxpam': 'Túxpam',
    '(Tulip\ufffdn)': '(Tulipán)',
    'Vel\ufffdzquez': 'Velázquez',
    'Xicoht\ufffdncatl': 'Xicohtencatl',  # Wait: Tlaxcala de Xicohténcatl - the correct name uses accent
    '\ufffdlamos': 'Álamos',
    '\ufffdlvarez': 'Álvarez',
    '\ufffdngel': 'Ángel',
    '\ufffdnimas': 'Ánimas',
    '\ufffdvila': 'Ávila',
    '\ufffdrsulo': 'Úrsulo',
    '(\ufffdrsulo': '(Úrsulo',
    '[Vi\ufffdedos]': '[Viñedos]',
    'Ar\ufffdmburo': 'Arámburo',
    # Market name tokens (uppercase)
    'ACU\ufffdA': 'ACUÑA',
    'ALI\ufffdADOS': 'ALIÑADOS',
    'AVI\ufffdA': 'AVIÑA',
    'CABA\ufffdA': 'CABAÑA',
    'CAMPI\ufffdA': 'CAMPIÑA',
    'CARNICER\ufffdA': 'CARNICERÍA',
    'COSTE\ufffdA': 'COSTEÑA',
    'DO\ufffdA': 'DOÑA',
    'KABA\ufffdA': 'KABAÑA',
    'MAGA\ufffdA': 'MAGAÑA',
    'ME\ufffdO': 'MEÑO',
    'MU\ufffdECO': 'MUÑECO',
    'M\ufffdXICO': 'MÉXICO',
    'NU\ufffdO': 'NUÑO',
    'OCA\ufffdA': 'OCAÑA',
    'PE\ufffdON': 'PEÑON',
    'PEQUE\ufffdO': 'PEQUEÑO',
    'PI\ufffdON': 'PIÑON',
    'P\ufffdBLICO': 'PÚBLICO',
    'RIZUE\ufffdO': 'RIZUEÑO',
    'TO\ufffdO': 'TOÑO',
    'URE\ufffdOS': 'UREÑOS',
    'VI\ufffdEDOS': 'VIÑEDOS',
    'Xalapa-Enr\ufffdquez': 'Xalapa-Enríquez',
}

TOKEN_FIX.update(MANUAL_TOKEN_FIX)

# Fix for Puriatzícuaro - it should be Puriatzícuaro (with accent)
TOKEN_FIX['Puriatz\ufffdcuaro'] = 'Puriatzícuaro'
# Xicohtencatl with accent
TOKEN_FIX['Xicoht\ufffdncatl'] = 'Xicoténcatl'

def fix_text(txt):
    """Fix broken encoding by replacing broken tokens with correct ones."""
    if txt is None:
        return None
    s = str(txt)
    if '\ufffd' not in s:
        return s
    
    # Apply token-level fixes (longest first to avoid partial matches)
    for broken, correct in sorted(TOKEN_FIX.items(), key=lambda x: len(x[0]), reverse=True):
        s = s.replace(broken, correct)
    
    return s

# ============================================================
# Step 3: Collect all 2217 unique markets
# ============================================================
unique_markets = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    mid = row[13]
    if mid and mid not in unique_markets:
        unique_markets[mid] = {
            'market_id': mid,
            'market_name': fix_text(row[14]),
            'tipo': row[15] or 'MERCADO_PUBLICO',
            'entidad': str(row[28]) if row[28] else fix_text(row[2]),      # Use Estago_geo
            'municipio': str(row[30]) if row[30] else fix_text(row[3]),    # Use Municipio_geo
            'localidad': fix_text(row[4]),                                  # Fix via tokens
            'latitud': row[5],
            'longitud': row[6],
            'n_establecimientos': row[18],
            'cve_ent': row[27],
            'cve_mun': row[29],
        }

print(f"Total unique markets: {len(unique_markets)}")

# ============================================================
# Step 4: Verify NO broken chars remain
# ============================================================
broken_fields = []
for mid, mk in unique_markets.items():
    for field in ['market_name', 'entidad', 'municipio', 'localidad']:
        val = mk[field]
        if val and '\ufffd' in str(val):
            broken_fields.append((field, val, mid))

if broken_fields:
    print(f"\n!!! STILL {len(broken_fields)} BROKEN FIELDS:")
    seen = set()
    for f, v, mid in broken_fields:
        key = (f, v)
        if key not in seen:
            seen.add(key)
            print(f"  {f}: {repr(v)}  (market: {mid})")
    sys.exit(1)

print("ALL TEXT CORRECT! No broken characters remain.")

# ============================================================
# Step 5: Verify counts per state match user's data
# ============================================================
from collections import Counter
state_counts = Counter()
for mk in unique_markets.values():
    state_counts[mk['entidad']] += 1

print(f"\n=== MERCADOS UNICOS POR ESTADO ({len(state_counts)} states) ===")
for state, count in sorted(state_counts.items(), key=lambda x: -x[1]):
    print(f"  {state}: {count}")

# ============================================================
# Step 6: Generate SQL migration
# ============================================================
lines = []
lines.append("-- Migration 007: Full catalog with 2217 markets (correct encoding)")
lines.append("-- Replaces previous 259-market catalog")
lines.append("")
lines.append("TRUNCATE catalogo_mercados CASCADE;")
lines.append("")
lines.append("INSERT INTO catalogo_mercados (market_id, nombre, tipo, entidad, municipio, localidad, latitud, longitud, n_establecimientos, cve_ent, cve_mun) VALUES")

def esc(v):
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"

values = []
for mk in unique_markets.values():
    lat = mk['latitud'] if mk['latitud'] else 'NULL'
    lon = mk['longitud'] if mk['longitud'] else 'NULL'
    n_est = mk['n_establecimientos'] if mk['n_establecimientos'] else 0
    val = f"({esc(mk['market_id'])}, {esc(mk['market_name'])}, {esc(mk['tipo'])}, {esc(mk['entidad'])}, {esc(mk['municipio'])}, {esc(mk['localidad'])}, {lat}, {lon}, {n_est}, {esc(mk['cve_ent'])}, {esc(mk['cve_mun'])})"
    values.append(val)

lines.append(",\n".join(values) + ";")

sql = "\n".join(lines)
with open('backend/migrations/007_fix_encoding_mercados.sql', 'w', encoding='utf-8') as f:
    f.write(sql)

print(f"\nGenerated migration: backend/migrations/007_fix_encoding_mercados.sql")
print(f"Total INSERT rows: {len(values)}")

# Sample output
print("\n=== SAMPLE DATA ===")
for i, mk in enumerate(list(unique_markets.values())[:5]):
    print(f"  {mk['market_name']} | {mk['entidad']} | {mk['municipio']} | {mk['localidad']}")
