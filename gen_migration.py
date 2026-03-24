import openpyxl
import json

wb = openpyxl.load_workbook('mercados_estado_mun.xlsx')
ws = wb['capa_unida']

catalogo_limpio = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    estatus = row[16]
    market_id = row[13]
    market_name = row[14]
    if market_id and estatus == 'CATALOGO_LIMPIO':
        if market_id not in catalogo_limpio:
            catalogo_limpio[market_id] = {
                'market_id': market_id,
                'market_name': market_name,
                'tipo': row[15],
                'entidad': row[2],
                'municipio': row[3],
                'localidad': row[4],
                'latitud': row[5],
                'longitud': row[6],
                'nom_CenCom': row[12],
                'tipoCenCom': row[11],
                'n_establecimientos': row[18],
                'cve_ent': row[27],
                'cve_mun': row[29],
            }

lines = []
lines.append("-- Migration 006: Catalogo de mercados pre-poblado desde Excel")
lines.append("-- 259 mercados con estatus CATALOGO_LIMPIO")
lines.append("")
lines.append("-- 1) Crear tabla catalogo_mercados")
lines.append("CREATE TABLE IF NOT EXISTS catalogo_mercados (")
lines.append("    id SERIAL PRIMARY KEY,")
lines.append("    market_id VARCHAR(30) UNIQUE NOT NULL,")
lines.append("    nombre VARCHAR(200) NOT NULL,")
lines.append("    tipo VARCHAR(50) NOT NULL,")
lines.append("    entidad VARCHAR(100) NOT NULL,")
lines.append("    municipio VARCHAR(100) NOT NULL,")
lines.append("    localidad VARCHAR(150),")
lines.append("    latitud DOUBLE PRECISION,")
lines.append("    longitud DOUBLE PRECISION,")
lines.append("    n_establecimientos INTEGER DEFAULT 0,")
lines.append("    cve_ent VARCHAR(5),")
lines.append("    cve_mun VARCHAR(5)")
lines.append(");")
lines.append("")
lines.append("-- 2) Modificar tabla mercados: agregar referencia al catalogo")
lines.append("ALTER TABLE mercados ADD COLUMN IF NOT EXISTS catalogo_mercado_id INTEGER REFERENCES catalogo_mercados(id);")
lines.append("ALTER TABLE mercados ADD COLUMN IF NOT EXISTS latitud DOUBLE PRECISION;")
lines.append("ALTER TABLE mercados ADD COLUMN IF NOT EXISTS longitud DOUBLE PRECISION;")
lines.append("")
lines.append("-- 3) Insertar los 259 mercados del catalogo")
lines.append("INSERT INTO catalogo_mercados (market_id, nombre, tipo, entidad, municipio, localidad, latitud, longitud, n_establecimientos, cve_ent, cve_mun) VALUES")

values = []
for mk in catalogo_limpio.values():
    def esc(v):
        if v is None:
            return "NULL"
        return "'" + str(v).replace("'", "''") + "'"
    
    lat = mk['latitud'] if mk['latitud'] else 'NULL'
    lon = mk['longitud'] if mk['longitud'] else 'NULL'
    n_est = mk['n_establecimientos'] if mk['n_establecimientos'] else 0
    
    val = f"({esc(mk['market_id'])}, {esc(mk['market_name'])}, {esc(mk['tipo'])}, {esc(mk['entidad'])}, {esc(mk['municipio'])}, {esc(mk['localidad'])}, {lat}, {lon}, {n_est}, {esc(mk['cve_ent'])}, {esc(mk['cve_mun'])})"
    values.append(val)

lines.append(",\n".join(values) + ";")
lines.append("")
lines.append("-- 4) Crear indices para busqueda rapida")
lines.append("CREATE INDEX IF NOT EXISTS idx_cat_mercados_entidad ON catalogo_mercados(entidad);")
lines.append("CREATE INDEX IF NOT EXISTS idx_cat_mercados_municipio ON catalogo_mercados(municipio);")
lines.append("CREATE INDEX IF NOT EXISTS idx_cat_mercados_nombre ON catalogo_mercados(nombre);")
lines.append("CREATE INDEX IF NOT EXISTS idx_cat_mercados_tipo ON catalogo_mercados(tipo);")

sql = "\n".join(lines)
with open('backend/migrations/006_catalogo_mercados.sql', 'w', encoding='utf-8') as f:
    f.write(sql)

print(f"Generated migration with {len(catalogo_limpio)} markets")
