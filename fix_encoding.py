import openpyxl

wb = openpyxl.load_workbook('mercados_estado_mun.xlsx')
ws = wb['capa_unida']

# Mapa completo de correcciones
FIX = {
    # Entidades
    'Ciudad de M\ufffdxico': 'Ciudad de México',
    'M\ufffdxico': 'México',
    'Michoac\ufffdn de Ocampo': 'Michoacán de Ocampo',
    'Nuevo Le\ufffdn': 'Nuevo León',
    'Quer\ufffdtaro': 'Querétaro',
    'San Luis Potos\ufffd': 'San Luis Potosí',
    'Yucat\ufffdn': 'Yucatán',
    # Municipios
    'Acapulco de Ju\ufffdrez': 'Acapulco de Juárez',
    'Ac\ufffdmbaro': 'Acámbaro',
    'Apatzing\ufffdn': 'Apatzingán',
    'Benito Ju\ufffdrez': 'Benito Juárez',
    'Cocotitl\ufffdn': 'Cocotitlán',
    'Comit\ufffdn de Dom\ufffdnguez': 'Comitán de Domínguez',
    'Cosal\ufffd': 'Cosalá',
    'Cuauht\ufffdmoc': 'Cuauhtémoc',
    'Cuautitl\ufffdn': 'Cuautitlán',
    'Cuitl\ufffdhuac': 'Cuitláhuac',
    'Culiac\ufffdn': 'Culiacán',
    'C\ufffdrdenas': 'Cárdenas',
    'C\ufffdrdoba': 'Córdoba',
    'El Marqu\ufffds': 'El Marqués',
    'Iguala de la Independencia': 'Iguala de la Independencia',
    'Ju\ufffdrez': 'Juárez',
    'Kanas\ufffdn': 'Kanasín',
    'Le\ufffdn': 'León',
    'Minatitl\ufffdn': 'Minatitlán',
    'M\ufffdrida': 'Mérida',
    'Nezahualc\ufffdyotl': 'Nezahualcóyotl',
    'Oaxaca de Ju\ufffdrez': 'Oaxaca de Juárez',
    'Ocotl\ufffdn': 'Ocotlán',
    'Santiago Miahuatl\ufffdn': 'Santiago Miahuatlán',
    'Santo Domingo Tonal\ufffd': 'Santo Domingo Tonalá',
    'Soledad de Graciano S\ufffdnchez': 'Soledad de Graciano Sánchez',
    'Soyaniquilpan de Ju\ufffdrez': 'Soyaniquilpan de Juárez',
    'Tac\ufffdmbaro': 'Tacámbaro',
    'Taxco de Alarc\ufffdn': 'Taxco de Alarcón',
    'Tecolotl\ufffdn': 'Tecolotlán',
    'Tecom\ufffdn': 'Tecomán',
    'Tec\ufffdmac': 'Tecámac',
    'Tomatl\ufffdn': 'Tomatlán',
    'Tultitl\ufffdn': 'Tultitlán',
    'Villa Ju\ufffdrez': 'Villa Juárez',
    'Villa de Tamazul\ufffdpam del Progreso': 'Villa de Tamazulápam del Progreso',
    'Xicot\ufffdncatl': 'Xicoténcatl',
    # Localidades
    'Apatzing\ufffdn de la Constituci\ufffdn': 'Apatzingán de la Constitución',
    'Canc\ufffdn': 'Cancún',
    'Ciudad Coahuila (Kil\ufffdmetro Cincuenta y Siete)': 'Ciudad Coahuila (Kilómetro Cincuenta y Siete)',
    'Ciudad Nezahualc\ufffdyotl': 'Ciudad Nezahualcóyotl',
    'Culiac\ufffdn Rosales': 'Culiacán Rosales',
    'General Miguel Alem\ufffdn (Potrero Nuevo)': 'General Miguel Alemán (Potrero Nuevo)',
    'Jerez de Garc\ufffda Salinas': 'Jerez de García Salinas',
    'Le\ufffdn de los Aldama': 'León de los Aldama',
    'Ocuiltzapotl\ufffdn': 'Ocuiltzapotlán',
    'Par\ufffdcuaro': 'Parácuaro',
    'Paseos del Marqu\ufffds': 'Paseos del Marqués',
    'San Jos\ufffd Guadalupe Otzacatipan': 'San José Guadalupe Otzacatipan',
    'San Jos\ufffd Monte Chiquito': 'San José Monte Chiquito',
    'San Juan del R\ufffdo': 'San Juan del Río',
    'San Mart\ufffdn Cuautlalpan': 'San Martín Cuautlalpan',
    'San Sebasti\ufffdn Zinacatepec': 'San Sebastián Zinacatepec',
    'Tac\ufffdmbaro de Codallos': 'Tacámbaro de Codallos',
    'Tapachula de C\ufffdrdova y Ord\ufffd\ufffdez': 'Tapachula de Córdova y Ordóñez',
    'Santiago de Quer\ufffdtaro': 'Santiago de Querétaro',
    'San Luis Potos\ufffd': 'San Luis Potosí',
    'Taxco de Alarc\ufffdn': 'Taxco de Alarcón',
    'Tempoal de S\ufffdnchez': 'Tempoal de Sánchez',
    'Tlalmanalco de Vel\ufffdzquez': 'Tlalmanalco de Velázquez',
    'Amecameca de Ju\ufffdrez': 'Amecameca de Juárez',
    'Xalapa-Enr\ufffdquez': 'Xalapa-Enríquez',
    '\ufffdngel R. Cabada': 'Ángel R. Cabada',
    # Market names
    'CENTRAL DE ABASTO DE LA CIUDAD DE M\ufffdXICO': 'CENTRAL DE ABASTO DE LA CIUDAD DE MÉXICO',
    'MERCADO GREGORIO MENDEZ MAGA\ufffdA': 'MERCADO GREGORIO MENDEZ MAGAÑA',
    'MERCADO P\ufffdBLICO SIN NOMBRE': 'MERCADO PÚBLICO SIN NOMBRE',
}

def fix_text(txt):
    if txt is None:
        return None
    s = str(txt)
    # Apply longest replacements first to avoid partial matches
    for bad, good in sorted(FIX.items(), key=lambda x: len(x[0]), reverse=True):
        s = s.replace(bad, good)
    return s

catalogo_limpio = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    estatus = row[16]
    market_id = row[13]
    if market_id and estatus == 'CATALOGO_LIMPIO':
        if market_id not in catalogo_limpio:
            catalogo_limpio[market_id] = {
                'market_id': market_id,
                'market_name': fix_text(row[14]),
                'tipo': row[15],
                'entidad': fix_text(row[2]),
                'municipio': fix_text(row[3]),
                'localidad': fix_text(row[4]),
                'latitud': row[5],
                'longitud': row[6],
                'n_establecimientos': row[18],
                'cve_ent': row[27],
                'cve_mun': row[29],
            }

# Verify no broken chars remain
broken = []
for mk in catalogo_limpio.values():
    for field in ['market_name', 'entidad', 'municipio', 'localidad']:
        val = mk[field]
        if val and '\ufffd' in val:
            broken.append((field, val))

if broken:
    print("!!! STILL BROKEN:")
    for f, v in sorted(set(broken)):
        print(f"  {f}: {repr(v)}")
    exit(1)

print("All text corrected! Generating migration...")

# Generate UPDATE SQL
lines = []
lines.append("-- Migration 007: Fix encoding in catalogo_mercados")
lines.append("-- Corrects broken Unicode characters in names")
lines.append("")

# Easier: TRUNCATE and re-insert with correct data
lines.append("TRUNCATE catalogo_mercados CASCADE;")
lines.append("")
lines.append("INSERT INTO catalogo_mercados (market_id, nombre, tipo, entidad, municipio, localidad, latitud, longitud, n_establecimientos, cve_ent, cve_mun) VALUES")

values = []
for mk in catalogo_limpio.values():
    def esc(v):
        if v is None:
            return "NULL"
        return "'" + str(v).replace("'", "''") + "'"
    lat = mk['latitud'] if mk['latitud'] else 'NULL'
    lon = mk['longitud'] if mk['longitud'] else 'NULL'
    n_est = mk['n_establecimientos'] if mk['n_establecimientos'] else 0
    val = f"({esc(mk['market_id'])}, {esc(mk['market_name'])}, {esc(mk['tipo'])}, {esc(mk['entidad'])}, {esc(mk['municipio'])}, {esc(mk['localidad'])}, {lat}, {lon}, {n_est}, {esc(mk['cve_ent'])}, {esc(mk['cve_mun'])})"
    values.append(val)

lines.append(",\n".join(values) + ";")

sql = "\n".join(lines)
with open('backend/migrations/007_fix_encoding_mercados.sql', 'w', encoding='utf-8') as f:
    f.write(sql)

print(f"Generated migration with {len(catalogo_limpio)} corrected markets")

# Show sample corrected data
print("\n=== SAMPLE CORRECTED DATA ===")
for i, mk in enumerate(list(catalogo_limpio.values())[:10]):
    print(f"  {mk['market_name']} | {mk['entidad']} | {mk['municipio']} | {mk['localidad']}")
