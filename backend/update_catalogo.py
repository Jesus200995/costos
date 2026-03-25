"""
Script para actualizar catálogos (categorias, subcategorias, productos, unidades)
desde catalogo.xlsx. Se ejecuta directamente en el servidor.
"""
import openpyxl
import psycopg2
import psycopg2.extras

DB = dict(dbname="costos", user="jesus", password="2025", host="localhost")

def main():
    wb = openpyxl.load_workbook("/var/www/costos/catalogo.xlsx")

    # ── Parse Excel ──────────────────────────────────────────────
    ws_prod = wb["Catalogo_Productos"]
    ws_units = wb["Unidades"]

    # Unit ID → label mapping
    unit_map = {}
    for r in range(2, ws_units.max_row + 1):
        uid = ws_units.cell(r, 1).value
        label = ws_units.cell(r, 2).value
        if uid and label:
            unit_map[uid.strip()] = label.strip()

    # Categories, subcategories, products, units per subcategoria
    categorias = {}      # id -> nombre
    subcategorias = {}   # id -> (nombre, categoria_id)
    productos = []       # (nombre, subcategoria_id)
    sub_menudeo = {}     # subcategoria_id -> set of unit labels
    sub_mayoreo = {}     # subcategoria_id -> set of unit labels

    for r in range(2, ws_prod.max_row + 1):
        c1id = ws_prod.cell(r, 1).value
        c1name = ws_prod.cell(r, 2).value
        sid = ws_prod.cell(r, 3).value
        sname = ws_prod.cell(r, 4).value
        pid = ws_prod.cell(r, 5).value
        pname = ws_prod.cell(r, 6).value
        men_ids = ws_prod.cell(r, 11).value
        may_ids = ws_prod.cell(r, 13).value

        if c1id and c1name:
            categorias[c1id.strip()] = c1name.strip()
        if sid and sname:
            sname_clean = sname.strip()
            # "Otros / Revisar" → "Otros"
            if "otros" in sname_clean.lower() and "revisar" in sname_clean.lower():
                sname_clean = "Otros"
            subcategorias[sid.strip()] = (sname_clean, c1id.strip())
        if pid and pname and sid:
            productos.append((pname.strip(), sid.strip()))

        if sid and men_ids:
            sid_s = sid.strip()
            if sid_s not in sub_menudeo:
                sub_menudeo[sid_s] = set()
            for u in str(men_ids).split("|"):
                u = u.strip()
                if u and u in unit_map:
                    sub_menudeo[sid_s].add(unit_map[u])

        if sid and may_ids:
            sid_s = sid.strip()
            if sid_s not in sub_mayoreo:
                sub_mayoreo[sid_s] = set()
            for u in str(may_ids).split("|"):
                u = u.strip()
                if u and u in unit_map:
                    sub_mayoreo[sid_s].add(unit_map[u])

    print(f"Parsed: {len(categorias)} cats, {len(subcategorias)} subs, "
          f"{len(productos)} prods, {len(unit_map)} unit types")

    # ── Apply to DB ──────────────────────────────────────────────
    conn = psycopg2.connect(**DB)
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        # 1. Upsert categorias
        for cid, cname in categorias.items():
            cur.execute(
                """INSERT INTO categorias (id, nombre)
                   VALUES (%s, %s)
                   ON CONFLICT (id) DO UPDATE SET nombre = EXCLUDED.nombre""",
                (cid, cname),
            )
        print(f"✓ Categorías: {len(categorias)}")

        # 2. Upsert subcategorias
        for sid, (sname, cid) in subcategorias.items():
            cur.execute(
                """INSERT INTO subcategorias (id, categoria_id, nombre)
                   VALUES (%s, %s, %s)
                   ON CONFLICT (id) DO UPDATE SET
                     categoria_id = EXCLUDED.categoria_id,
                     nombre = EXCLUDED.nombre""",
                (sid, cid, sname),
            )
        print(f"✓ Subcategorías: {len(subcategorias)}")

        # 3. Products: preserve existing IDs where name+subcategoria match
        # Get existing products
        cur.execute("SELECT id, subcategoria_id, nombre FROM productos")
        existing_prods = {(r["subcategoria_id"], r["nombre"]): r["id"] for r in cur.fetchall()}

        # Also get products referenced in detalle_precios (cannot delete)
        cur.execute("SELECT DISTINCT producto_id FROM detalle_precios")
        used_product_ids = {r["producto_id"] for r in cur.fetchall()}

        new_prod_keys = set()
        inserted = 0
        kept = 0
        for pname, sid in productos:
            new_prod_keys.add((sid, pname))
            if (sid, pname) in existing_prods:
                kept += 1
            else:
                cur.execute(
                    """INSERT INTO productos (subcategoria_id, nombre)
                       VALUES (%s, %s)
                       ON CONFLICT (subcategoria_id, nombre) DO NOTHING""",
                    (sid, pname),
                )
                inserted += 1

        # Delete products that are no longer in catalog (if not used in prices)
        deleted = 0
        for (sid, pname), pid in existing_prods.items():
            if (sid, pname) not in new_prod_keys:
                if pid in used_product_ids:
                    print(f"  ⚠ Producto {pid} '{pname}' tiene precios, no se elimina")
                else:
                    cur.execute("DELETE FROM productos WHERE id = %s", (pid,))
                    deleted += 1

        print(f"✓ Productos: {kept} existentes, {inserted} nuevos, {deleted} eliminados")

        # 4. Units: clear and reinsert
        # Check which units are used in detalle_precios
        cur.execute("SELECT DISTINCT unidad FROM detalle_precios")
        used_units = {r["unidad"] for r in cur.fetchall()}

        cur.execute("DELETE FROM unidades_subcategoria")
        total_units = 0

        for sid in subcategorias:
            menudeo_labels = sub_menudeo.get(sid, set())
            mayoreo_labels = sub_mayoreo.get(sid, set())

            for label in menudeo_labels:
                cur.execute(
                    """INSERT INTO unidades_subcategoria (subcategoria_id, nombre, tipo)
                       VALUES (%s, %s, 'MENUDEO')
                       ON CONFLICT (subcategoria_id, nombre, tipo) DO NOTHING""",
                    (sid, label),
                )
                total_units += 1

            for label in mayoreo_labels:
                cur.execute(
                    """INSERT INTO unidades_subcategoria (subcategoria_id, nombre, tipo)
                       VALUES (%s, %s, 'MAYOREO')
                       ON CONFLICT (subcategoria_id, nombre, tipo) DO NOTHING""",
                    (sid, label),
                )
                total_units += 1

        print(f"✓ Unidades: {total_units} registros insertados")

        conn.commit()
        print("\n✅ Catálogo actualizado exitosamente")

        # Verification
        cur.execute("SELECT COUNT(*) as c FROM categorias")
        print(f"   Categorías: {cur.fetchone()['c']}")
        cur.execute("SELECT COUNT(*) as c FROM subcategorias")
        print(f"   Subcategorías: {cur.fetchone()['c']}")
        cur.execute("SELECT COUNT(*) as c FROM productos")
        print(f"   Productos: {cur.fetchone()['c']}")
        cur.execute("SELECT COUNT(*) as c FROM unidades_subcategoria")
        print(f"   Unidades: {cur.fetchone()['c']}")

    except Exception as e:
        conn.rollback()
        print(f"❌ Error: {e}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
