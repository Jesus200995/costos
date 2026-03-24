import openpyxl

wb = openpyxl.load_workbook('mercados_estado_mun.xlsx')
ws = wb['capa_unida']

estatus_vals = set()
catalogo_limpio = {}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    estatus = row[16]
    market_id = row[13]
    market_name = row[14]
    if estatus:
        estatus_vals.add(estatus)
    if market_id and estatus == 'CATALOGO_LIMPIO':
        if market_id not in catalogo_limpio:
            catalogo_limpio[market_id] = {
                'market_name': market_name,
                'tipo': row[15],
                'entidad': row[2],
                'municipio': row[3],
                'localidad': row[4],
                'latitud': row[5],
                'longitud': row[6],
                'nom_CenCom': row[12],
                'tipoCenCom': row[11],
                'n_establecimientos': row[18],
                'CVE_ENT': row[27],
                'CVE_MUN': row[29],
            }

print(f"Estatus values: {estatus_vals}")
print(f"Total CATALOGO_LIMPIO unique markets: {len(catalogo_limpio)}")
print()

estados = set()
for mk in catalogo_limpio.values():
    estados.add(mk['entidad'])
print(f"Estados representados: {len(estados)}")
print(sorted(estados))
print()

tipos = {}
for mk in catalogo_limpio.values():
    t = mk['tipo']
    tipos[t] = tipos.get(t, 0) + 1
print(f"Tipos de mercado: {tipos}")
print()

for i, (mid, m) in enumerate(list(catalogo_limpio.items())[:20]):
    print(f"{mid} | {m['market_name']} | {m['entidad']} | {m['municipio']} | {m['localidad']} | {m['latitud']},{m['longitud']} | {m['tipo']} | n_est:{m['n_establecimientos']}")
