"""
Collect ALL unique broken strings from the Excel to build a comprehensive fix map.
"""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('mercados_estado_mun.xlsx')
ws = wb['capa_unida']

# Gather all unique market entries
unique_markets = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    mid = row[13]
    if mid and mid not in unique_markets:
        unique_markets[mid] = row

# Collect ALL broken strings by field
from collections import defaultdict
broken = defaultdict(set)
for mid, r in unique_markets.items():
    for idx, field in [(2, 'entidad'), (3, 'municipio'), (4, 'localidad'), (14, 'market_name')]:
        val = r[idx]
        if val and '\ufffd' in str(val):
            broken[field].add(str(val))

for field in ['entidad', 'municipio', 'localidad', 'market_name']:
    vals = sorted(broken.get(field, []))
    print(f"\n=== {field.upper()} ({len(vals)} broken) ===")
    for v in vals:
        print(f"  {repr(v)}")
