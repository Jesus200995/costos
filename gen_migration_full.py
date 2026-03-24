import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('mercados_estado_mun.xlsx')
ws = wb['capa_unida']

# First pass: collect ALL rows and understand column layout
headers = [cell.value for cell in ws[1]]
print("=== COLUMNS ===")
for i, h in enumerate(headers):
    print(f"  [{i}] {h}")

# Gather all unique market_id entries
all_rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    all_rows.append(row)

print(f"\nTotal rows: {len(all_rows)}")

# Count records with market_id
with_market = [r for r in all_rows if r[13]]  # market_id at col 13
without_market = [r for r in all_rows if not r[13]]
print(f"Rows with market_id: {len(with_market)}")
print(f"Rows without market_id: {len(without_market)}")

# Unique markets by market_id
unique_markets = {}
for r in with_market:
    mid = r[13]
    if mid not in unique_markets:
        unique_markets[mid] = r

print(f"Unique markets (by market_id): {len(unique_markets)}")

# Check all unique estatus values
estatus_vals = set()
for r in all_rows:
    estatus_vals.add(r[16])
print(f"\nUnique estatus values: {estatus_vals}")

# Count by estatus
from collections import Counter
estatus_counts = Counter(r[16] for r in with_market)
print(f"Markets by estatus: {estatus_counts}")

# Unique markets by estatus
estatus_unique = Counter()
for mid, r in unique_markets.items():
    estatus_unique[r[16]] += 1
print(f"Unique markets by estatus: {estatus_unique}")

# Check tipo values
tipo_counts = Counter(r[15] for r in unique_markets.values())
print(f"\nTypes: {tipo_counts}")

# Count unique markets per state
state_counts = Counter()
for mid, r in unique_markets.items():
    state_counts[str(r[2])] += 1

print("\n=== MERCADOS UNICOS POR ESTADO ===")
for state, count in sorted(state_counts.items(), key=lambda x: -x[1]):
    print(f"  {state}: {count}")

# Check for broken chars
broken_count = 0
for mid, r in unique_markets.items():
    for idx in [2, 3, 4, 14]:  # entidad, municipio, localidad, market_name
        val = r[idx]
        if val and '\ufffd' in str(val):
            broken_count += 1

print(f"\nFields with broken encoding: {broken_count}")
